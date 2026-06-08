"""Export the full training plan to Excel."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = Workbook()

# --- Colors ---
PHASE_FILLS = {
    'Base': PatternFill('solid', start_color='E0F7FA'),
    'Build': PatternFill('solid', start_color='EDE7F6'),
    'Peak': PatternFill('solid', start_color='FBE9E7'),
    'Taper': PatternFill('solid', start_color='E8F5E9'),
    'Race': PatternFill('solid', start_color='FCE4EC'),
}
HEADER_FILL = PatternFill('solid', start_color='1A2340')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
DATA_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
RECOVERY_FILL = PatternFill('solid', start_color='F5F5F5')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
)
PLAN_START = datetime(2026, 6, 9)

PLAN = [
    (1,'9 jun','Base',45,24,'Easy volume opbouw + 2x krachttraining','Aerobe basis',False),
    (2,'16 jun','Base',48,26,'Tempo run 8 km + trail run','Volume opbouwen',False),
    (3,'23 jun','Base',52,28,'Fartlek 10 km + easy doubles','Long run verlengen',False),
    (4,'30 jun','Base',42,22,'Recovery week — alle runs easy pace','Herstelweek',True),
    (5,'7 jul','Base',55,30,'Tempo 10 km + hill repeats + B2B intro (30+12)','B2B intro',False),
    (6,'14 jul','Base',58,30,'Threshold run + long run op trails','Trail gewenning',False),
    (7,'21 jul','Build',62,32,'Back-to-back weekend (32+15) + tempo','B2B opbouwen',False),
    (8,'28 jul','Build',65,30,'Hitte-run middaguur + sauna protocol start','Hitte-adaptatie',False),
    (9,'4 aug','Build',50,22,'Recovery week — easy + cross-training','Herstelweek',True),
    (10,'11 aug','Build',68,35,'B2B (35+20) + zandtraining','Vermoeidheid tolerantie',False),
    (11,'18 aug','Build',72,35,'Voedingsstrategie oefenen op long run','Voeding testen',False),
    (12,'25 aug','Build',75,38,'B2B (38+22) + mentale training','Piekaanloop',False),
    (13,'1 sep','Peak',60,28,'Recovery + race gear testen','Herstelweek',True),
    (14,'8 sep','Peak',75,38,'Race simulatie #1: 38+25 km B2B weekend','Simulatie 1',False),
    (15,'15 sep','Peak',80,40,'Piekweek: 40+25 B2B + midweek tempo','MAX VOLUME',False),
    (16,'22 sep','Peak',60,25,'Recovery — easy runs + sauna','Herstelweek',True),
    (17,'29 sep','Peak',78,38,'Race simulatie #2: 38+25 km B2B + race-voeding','Simulatie 2',False),
    (18,'6 okt','Peak',75,38,'Laatste grote week: 38+22 B2B + gear finaal','Laatste peak',False),
    (19,'13 okt','Taper',55,25,'Volume -30% — easy runs + strides','Afbouw start',False),
    (20,'20 okt','Taper',45,20,'Volume -45% — focus slaap & voeding','Dieper afbouwen',False),
    (21,'27 okt','Taper',35,16,'Lichte runs + strides. Packing.','Race voorbereiding',False),
    (22,'3 nov','Taper',25,10,'Shakeout runs. Mentale visualisatie.','Laatste rust',False),
    (23,'10 nov','Race',110,60,'Ma-Wo shakeout. Do rust. Vr 60 km + Za 50 km RACE!','RACE DAY!',False),
]

WEEKLY = {
    1:[('Ma','Rust','Rustdag. Foam rolling + IT-band oefeningen.',0),('Di','Easy','Easy run zone 2 — comfortabel tempo, focus op ademhaling.',8),('Wo','Tempo','Warm-up 2 km, dan 4 km aan tempo (~4:50/km), cool-down 2 km.',8),('Do','Kracht','Recovery run + krachttraining: squats, lunges, planks, clamshells.',6),('Vr','Easy','Easy run zone 2. Korter, gericht op herstel voor het weekend.',7),('Za','Long Run','Long run op gemakkelijk tempo. Begin voedingsstrategie te oefenen.',24),('Zo','Rust','Volledige rust of wandeling. Stretching en foam rolling.',0)],
    2:[('Ma','Rust','Rustdag. Foam rolling + stretching.',0),('Di','Easy','Easy run zone 2 — los en ontspannen.',8),('Wo','Tempo','Tempo run: 2 km warm-up, 6 km @ tempo, 2 km cool-down.',10),('Do','Recovery','Lichte recovery run + krachttraining (core + glutes).',6),('Vr','Easy','Easy run op gevoel. Trail als mogelijk.',8),('Za','Long Run','Long run, geleidelijk opbouwen. Probeer trail/onverharde paden.',26),('Zo','Rust','Rust. Lichte wandeling optioneel.',0)],
    3:[('Ma','Rust','Rustdag. IT-band preventie oefeningen.',0),('Di','Easy','Easy run zone 2.',8),('Wo','Fartlek','Fartlek: 10 km met 8x (1 min hard / 2 min easy).',10),('Do','Kracht','Recovery 6 km + krachttraining.',6),('Vr','Easy','Easy run, evt. als double (ochtend + avond 4 km).',8),('Za','Long Run','Langste run tot nu toe. Neem gels/voeding mee om te oefenen.',28),('Zo','Rust','Volledige rust.',0)],
    4:[('Ma','Rust','Recovery week — alles op laag tempo.',0),('Di','Easy','Easy run zone 1-2. Echt langzaam.',8),('Wo','Recovery','Heel lichte run of cross-training (fietsen/zwemmen).',6),('Do','Rust','Rustdag. Stretching en foam rolling.',0),('Vr','Easy','Easy run, geniet ervan.',8),('Za','Long Run','Matige long run — niet pushen. Recovery week.',22),('Zo','Rust','Volledige rust.',0)],
    5:[('Ma','Rust','Rustdag. Foam rolling.',0),('Di','Easy','Easy run zone 2.',8),('Wo','Tempo','Tempo run 10 km met 5 km @ threshold.',10),('Do','Heuvels','Hill repeats: 6-8x 2 min bergop, jog terug. Kracht voor Marokko.',7),('Vr','Easy','Easy run ter voorbereiding op back-to-back weekend.',6),('Za','Long Run','Long run 30 km — eerste keer 30+! Neem genoeg voeding.',30),('Zo','B2B','Back-to-back intro: run op vermoeide benen. Langzaam tempo.',12)],
    6:[('Ma','Rust','Rustdag na B2B weekend.',0),('Di','Easy','Easy run zone 2.',8),('Wo','Threshold','Threshold run: 3 km warm-up, 5 km @ drempel, 2 km cool-down.',10),('Do','Kracht','Recovery run + kracht. Focus glutes en core.',6),('Vr','Easy','Easy run op trail/onverharde paden.',8),('Za','Long Run','Long run op trails. Oefen met race-vest en voeding.',30),('Zo','Rust','Rust. Evalueer Base phase — klaar voor Build!',0)],
    7:[('Ma','Rust','Rustdag.',0),('Di','Easy','Easy run zone 2.',9),('Wo','Tempo','Tempo run 12 km met 6 km @ race effort.',12),('Do','Kracht','Recovery run + krachttraining.',7),('Vr','Easy','Easy run, voorbereiden op B2B weekend.',7),('Za','Long Run','Long run 32 km. Oefen met hitte (extra laag of middag).',32),('Zo','B2B','Back-to-back: 15 km op vermoeide benen. Focus mental toughness.',15)],
    8:[('Ma','Rust','Rustdag na B2B. Extra stretching.',0),('Di','Easy','Easy run zone 2. Middag als hitte-training.',10),('Wo','Tempo','Tempo run met progressief opbouwen.',11),('Do','Kracht','Recovery run + kracht. Sauna na training (15 min).',7),('Vr','Easy','Easy run. Start sauna protocol: 15 min post-run.',9),('Za','Long Run','Long run 30 km. Probeer op het warmste deel van de dag.',30),('Zo','Rust','Rust. Hydratatie focus.',0)],
    9:[('Ma','Rust','Recovery week — alle systemen laten herstellen.',0),('Di','Easy','Easy run zone 1-2. Heel relaxed.',8),('Wo','Cross','Cross-training: fietsen of zwemmen. Actief herstel.',0),('Do','Easy','Lichte easy run.',6),('Vr','Rust','Rustdag.',0),('Za','Long Run','Gematigde long run. Niet pushen.',22),('Zo','Rust','Volledige rust. Mentaal herladen voor Peak.',0)],
    10:[('Ma','Rust','Rustdag.',0),('Di','Easy','Easy run zone 2.',10),('Wo','Tempo','Tempo run 12 km. Oefen met gels onderweg.',12),('Do','Kracht','Recovery run + kracht + sauna.',7),('Vr','Easy','Easy run. Probeer op zand (strand) te lopen.',7),('Za','Long Run','35 km long run! Neem race-vest + volledige voeding mee.',35),('Zo','B2B','Back-to-back 20 km. Train mentale weerbaarheid.',20)],
    11:[('Ma','Rust','Rustdag na zwaar B2B weekend.',0),('Di','Easy','Easy run zone 2.',10),('Wo','Tempo','Tempo/threshold run. Voeding timen oefenen.',12),('Do','Kracht','Recovery run + krachttraining + sauna.',7),('Vr','Easy','Easy run voorbereidend op weekend.',8),('Za','Long Run','35 km met volledige race-voeding strategie. Alles testen!',35),('Zo','Rust','Rust. Evalueer voedingsplan.',0)],
    12:[('Ma','Rust','Rustdag. Laatste week van Build phase.',0),('Di','Easy','Easy run zone 2.',10),('Wo','Tempo','Tempo run 10 km. Mentale visualisatie oefenen.',10),('Do','Kracht','Recovery run + kracht.',7),('Vr','Easy','Easy run. Mentaal voorbereiden op B2B.',6),('Za','Long Run','38 km — langste training run! Alles op race-niveau.',38),('Zo','B2B','Back-to-back 22 km. Simuleer dag 2 gevoel.',22)],
    13:[('Ma','Rust','Recovery week — Peak phase begint.',0),('Di','Easy','Easy run zone 2.',8),('Wo','Recovery','Lichte run + test al je race gear.',6),('Do','Kracht','Recovery + kracht. Gear checklist maken.',6),('Vr','Easy','Easy run met race-schoenen.',8),('Za','Long Run','Gematigde long run 28 km. Test gamaschen + vest.',28),('Zo','Rust','Rust. Race gear evaluatie.',0)],
    14:[('Ma','Rust','Rustdag. Grote week ahead.',0),('Di','Easy','Easy run zone 2.',10),('Wo','Tempo','Tempo run 12 km aan race-effort.',12),('Do','Kracht','Recovery run + kracht + sauna.',7),('Vr','Easy','Easy run. Mentaal voorbereiden op simulatie.',7),('Za','Long Run','RACE SIMULATIE #1: 38 km met race-voeding en gear. Focus op pacing.',38),('Zo','B2B','Dag 2 simulatie: 25 km op vermoeide benen. B2B ervaring opdoen!',25)],
    15:[('Ma','Rust','Rust na simulatie. Extra slaap.',0),('Di','Easy','Easy run, heel relaxed.',8),('Wo','Tempo','Tempo run met progressie. Piekweek!',12),('Do','Kracht','Recovery + kracht.',7),('Vr','Easy','Easy run. Voorbereiden op MAX weekend.',7),('Za','Long Run','PIEKWEEK: 40 km — langste trainingsrun voor de race.',40),('Zo','B2B','Back-to-back 25 km. Push through! Mentale grenzen verleggen.',25)],
    16:[('Ma','Rust','Recovery week na piektraining.',0),('Di','Easy','Hele lichte easy run.',8),('Wo','Recovery','Recovery run of cross-training.',6),('Do','Rust','Rustdag. Sauna optioneel.',0),('Vr','Easy','Easy run zone 1-2.',8),('Za','Long Run','Gematigde long run. Niet pushen.',25),('Zo','Rust','Volledige rust.',0)],
    17:[('Ma','Rust','Rustdag. Grote simulatie komt eraan.',0),('Di','Easy','Easy run zone 2.',10),('Wo','Tempo','Tempo run 10 km. Scherp blijven.',10),('Do','Kracht','Recovery + lichte kracht.',7),('Vr','Easy','Easy run. Alles klaarzetten voor weekend.',7),('Za','Long Run','RACE SIMULATIE #2: 38 km met volledige race-setup en voeding.',38),('Zo','B2B','Dag 2: 25 km op vermoeide benen. Laatste grote B2B test!',25)],
    18:[('Ma','Rust','Rust na laatste grote simulatie.',0),('Di','Easy','Easy run zone 2.',10),('Wo','Tempo','Laatste tempo run voor taper. 8 km @ threshold.',10),('Do','Kracht','Recovery + laatste zware krachttraining.',7),('Vr','Easy','Easy run.',8),('Za','Long Run','Laatste lange run: 38 km. Geniet ervan!',38),('Zo','B2B','Laatste B2B: 22 km. Daarna begint taper.',22)],
    19:[('Ma','Rust','Start taper. Volume gaat omlaag.',0),('Di','Easy','Easy run zone 2. Geniet van het lichtere schema.',10),('Wo','Strides','Easy run 8 km met 6x strides (15 sec snel).',10),('Do','Kracht','Lichte run + lichte kracht. Geen zware sets meer.',6),('Vr','Easy','Easy run.',8),('Za','Long Run','Afgebouwde long run 25 km. Comfortabel tempo.',25),('Zo','Rust','Rust.',0)],
    20:[('Ma','Rust','Rustdag.',0),('Di','Easy','Easy run 8 km. Focus op slaapkwaliteit.',8),('Wo','Easy','Lichte run met strides.',8),('Do','Rust','Rustdag. Mentale visualisatie.',0),('Vr','Easy','Easy run. Voeding laden begint.',7),('Za','Long Run','Afgebouwde long run 20 km. Laatste langere effort.',20),('Zo','Rust','Rust. Packing lijst doorlopen.',0)],
    21:[('Ma','Rust','Rustdag.',0),('Di','Easy','Lichte run 7 km + strides.',7),('Wo','Easy','Easy run 6 km.',6),('Do','Rust','Rust. Koffer/tas pakken.',0),('Vr','Easy','Easy run 6 km. Check alle gear.',6),('Za','Long Run','Korte long run 16 km. Comfortabel.',16),('Zo','Rust','Rust. Alles klaarleggen.',0)],
    22:[('Ma','Rust','Laatste trainingsweek. Alleen shakeouts.',0),('Di','Shakeout','Shakeout run 5 km. Licht en los.',5),('Wo','Shakeout','Shakeout 5 km met paar strides.',5),('Do','Rust','Rust. Mentale voorbereiding. Visualiseer de finish.',0),('Vr','Shakeout','Laatste shakeout 3 km. Je bent er klaar voor!',3),('Za','Rust','Volledige rust. Carb loading. Vroeg naar bed.',0),('Zo','Rust','Rust. Reis naar Marokko als nodig.',0)],
    23:[('Ma','Shakeout','Aankomst Marokko. Lichte shakeout 3 km.',3),('Di','Shakeout','Shakeout 3 km. Acclimatiseer. Check routeinfo.',3),('Wo','Rust','Volledige rust. Hydrateer. Mentale focus.',0),('Do','Rust','Rust. Gear klaarzetten. Vroeg slapen.',0),('Vr','RACE D1','RACE DAG 1: 60 km door de Sahara. De langste etappe. YOU GOT THIS!',60),('Za','RACE D2','RACE DAG 2: 50 km. Finish strong. Je hebt hiervoor getraind!',50),('Zo','Feest!','Je hebt het gedaan! Herstel, vier het, en wees trots.',0)],
}

# ═══════ Sheet 1: Weekoverzicht ═══════
ws1 = wb.active
ws1.title = 'Weekoverzicht'
ws1.sheet_properties.tabColor = 'C2542D'

headers1 = ['Week', 'Startdatum', 'Fase', 'Doel km', 'Long Run km', 'Sleuteltrainingen', 'Focus', 'Herstelweek']
for c, h in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

ws1.row_dimensions[1].height = 30

for r, p in enumerate(PLAN, 2):
    wk, date, phase, km, lr, key, focus, recovery = p
    ws1.cell(row=r, column=1, value=wk).font = BOLD_FONT
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws1.cell(row=r, column=2, value=date).font = DATA_FONT
    ws1.cell(row=r, column=3, value=phase).font = DATA_FONT
    ws1.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws1.cell(row=r, column=4, value=km).font = BOLD_FONT
    ws1.cell(row=r, column=4).alignment = Alignment(horizontal='center')
    ws1.cell(row=r, column=5, value=lr).font = DATA_FONT
    ws1.cell(row=r, column=5).alignment = Alignment(horizontal='center')
    ws1.cell(row=r, column=6, value=key).font = DATA_FONT
    ws1.cell(row=r, column=6).alignment = Alignment(wrap_text=True)
    ws1.cell(row=r, column=7, value=focus).font = DATA_FONT
    ws1.cell(row=r, column=8, value='Ja' if recovery else '').font = DATA_FONT
    ws1.cell(row=r, column=8).alignment = Alignment(horizontal='center')

    fill = RECOVERY_FILL if recovery else PHASE_FILLS.get(phase)
    if fill:
        for c in range(1, 9):
            ws1.cell(row=r, column=c).fill = fill
    for c in range(1, 9):
        ws1.cell(row=r, column=c).border = THIN_BORDER

# Totals row
r_total = len(PLAN) + 2
ws1.cell(row=r_total, column=3, value='TOTAAL').font = BOLD_FONT
ws1.cell(row=r_total, column=4).font = BOLD_FONT
ws1.cell(row=r_total, column=4, value=f'=SUM(D2:D{r_total-1})')

col_widths1 = [7, 12, 8, 10, 13, 50, 22, 12]
for i, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f'A1:H{len(PLAN)+1}'

# ═══════ Sheet 2: Dagschema ═══════
ws2 = wb.create_sheet('Dagschema')
ws2.sheet_properties.tabColor = '22D3EE'

headers2 = ['Week', 'Fase', 'Datum', 'Dag', 'Type', 'Beschrijving', 'Km']
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

ws2.row_dimensions[1].height = 28
row = 2

for wk_idx, p in enumerate(PLAN):
    wk, date_str, phase, km, lr, key, focus, recovery = p
    days = WEEKLY[wk]
    week_start = PLAN_START + timedelta(days=(wk - 1) * 7)

    for d_idx, (dag, typ, desc, dkm) in enumerate(days):
        day_date = week_start + timedelta(days=d_idx)
        ws2.cell(row=row, column=1, value=wk).font = BOLD_FONT
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=2, value=phase).font = DATA_FONT
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=3, value=day_date.strftime('%Y-%m-%d')).font = DATA_FONT
        ws2.cell(row=row, column=4, value=dag).font = BOLD_FONT
        ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=5, value=typ).font = DATA_FONT
        ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws2.cell(row=row, column=6, value=desc).font = DATA_FONT
        ws2.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
        ws2.cell(row=row, column=7, value=dkm if dkm > 0 else '').font = BOLD_FONT
        ws2.cell(row=row, column=7).alignment = Alignment(horizontal='center')

        fill = RECOVERY_FILL if recovery else PHASE_FILLS.get(phase)
        if fill:
            for c in range(1, 8):
                ws2.cell(row=row, column=c).fill = fill
        for c in range(1, 8):
            ws2.cell(row=row, column=c).border = THIN_BORDER
        row += 1

col_widths2 = [7, 8, 12, 6, 12, 58, 7]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:G{row-1}'

out = r'C:\Users\fgeer\projects\ultra-x-dashboard\data\trainingsschema.xlsx'
wb.save(out)
print(f'Saved to {out} — {len(PLAN)} weeks, {row-2} day rows')
